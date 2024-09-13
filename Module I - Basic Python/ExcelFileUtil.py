import PlayerResults
import openpyxl

EXCEL_FILENAME = 'Estadistica.xlsx'
SHEET_NAME = 'Jugador y resultado'

def getFile():
    excel = openpyxl.load_workbook(EXCEL_FILENAME)
    return excel

def addGameResult(player, result):
    excel = getFile()
    hoja = excel[SHEET_NAME]
    hoja.append([player, result])
    excel.save(EXCEL_FILENAME)

def showInfo(nameIntroduced):
    playersStatistics = gatherInformation()
    if nameIntroduced != "":
        print(playersStatistics.get(nameIntroduced, 'No se ha encontrado'))
    else:
        print(playersStatistics)

def gatherInformation():
    players = {}
    excel = getFile()
    hoja = excel[SHEET_NAME]

    for i in range(3, hoja.max_row+1):
        playerName = hoja.cell(row=i, column=1).value
        playerResult = hoja.cell(row=i, column=2).value

        if players.get(playerName) == None:
            playerDictionary = {
                 playerName: []
            }
            playerDictionary[playerName].append(playerResult)
            players.update(playerDictionary)
        else:
            players.get(playerName).append(playerResult)
    
    return players